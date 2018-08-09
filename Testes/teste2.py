#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Fri Aug  3 06:55:40 2018
@author: C339182
"""

import openpyxl

path_to_file = "/home/gustavo/Documents/LPC & Co/SEC_CULT/EDITAIS/EDITAIS GERAIS Nº 2015 OFICIAL - Bárbara.xlsx"
texto_pivo = "NÚMERO DO EDITAL" #coluna 1, linha 5 do cabeçalho
sheet = "Plan1"

book = openpyxl.load_workbook(path_to_file)

planilha = book[sheet]

# No of written Rows in sheet
r = planilha.max_row

# No of written Columns in sheet
c = planilha.max_column

row_list = []

for linha in range(1, r+1):
    if planilha.cell(row=linha,column=1).value == "NÚMERO DO EDITAL":
        row_list.append(planilha.cell(row=linha,column=1).row)
        
# start = n_ed + 6
# end = n_ed[+1] - 6

col = 1
nomes = []

while col <= c+1:
    nomes.append(planilha.cell(row=7,column = col).value)
    print(nomes[-1])
    col += 3

nomes = nomes[1:-3]

Val = {}
temp = [] # triplas
i=1

for trip in nomes:
    Val[trip] = []
# divide and conquer, criar uma coluna gigante com todos os valores, sum(Val[trip][start-1:end-1])
while i <= r + 1:
    for trip in nomes:
        j = 1
        temp = []
        while j <= 3:
            coluna = j + 3*(nomes.index(trip)+1)
            if planilha.cell(row=i,column=coluna).value != None:
                temp.append(planilha.cell(row = i,column=coluna).value)
            else:
                temp.append(0)
            j += 1 # working
        Val[trip].append(temp) #working
    print("row")
    i +=1

def soma(valor, valor1, valor2, valor3):
    a=0
    if type(valor1) == str: 
        valor1 = valor1.replace("=","") # working
        valor1 = eval(valor1)
    elif type(valor2) == str: # checa se valor é string (padrão '=int+int')
        valor2 = valor2.replace("=","") # limpa o igual
        valor2 = eval(valor2) # faz a conta da string 'int+int'
    elif type(valor3) == str:
        valor3 = valor3.replace("=","")
        valor3 = eval(valor3)
    elif type(valor) == str:
        valor.replace("=","")
        valor = eval(valor)
    '''else:
        float(valor)
        float(valor1)
        float(valor2)
        float(valor3)'''
    a = valor+valor1+valor2+valor3 # retorna a soma
    return a

Hope = {}

for trip in nomes:
    Hope[trip] = []

for l in row_list:
    start = l + 5  # pega o começo do projeto
    try:
        ind = row_list.index(l)
        end = row_list[ind+1]-6 # pega o fim
    except IndexError:
        end = r-1
    for trip in nomes:
        valu = []
        som = 0
        values = Val[trip][start:end] # slice os valores dos projetos
        # print(trip)
        for valor in values:
            print(valor)
            # print(values.index(valor))
            som = soma(som, *valor) # working
            valu.append(som)
        Hope[trip].append(valu) # not exactly what I wanted, but so much better. cada lista no key é um proj

Fim = {}
for trip in nomes:
    Fim[trip] = []
    for proj in Hope[trip]:
        proje = sum(proj)
        Fim[trip].append(proje) # soma as listas de cada key e salva como coluna

import pandas as pd

df = pd.DataFrame(Fim) # exportando
df.to_excel('/home/gustavo/Documents/LPC & Co/SEC_CULT/condensado_2015.xlsx', index=False)
# Copiar o código em funções no jupyter pra poder rodar em loop pra puxar as info de vários arquivos